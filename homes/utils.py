from typing import Union
import os
from Config import settings
import openpyxl
from openpyxl.styles import Font,Alignment
from openpyxl.styles.colors import BLACK

def get_file_path(file_name:str) -> str:
    return os.path.join(settings.BASE_DIR,file_name) 



def write_galla_to_excel(file_name:str,sheet_name:str,gallas_data: Union[tuple,list]):
    file_path=get_file_path(file_name=file_name)
    book=openpyxl.open(filename=file_path)
    if not sheet_name in book.sheetnames:
        book.create_sheet(sheet_name)
    sheet=book[sheet_name]
    sheet.column_dimensions['A'].width=10
    sheet.column_dimensions['B'].width=11
    sheet.column_dimensions['C'].width=15
    sheet.column_dimensions['D'].width=30
    sheet.column_dimensions['E'].width=15
    sheet.column_dimensions['F'].width=11
    sheet.column_dimensions['G'].width=20
    sheet.column_dimensions['H'].width=20
    sheet.column_dimensions['I'].width=30
    sheet.column_dimensions['J'].width=11
    sheet.column_dimensions['K'].width=11
    sheet.column_dimensions['L'].width=11
    sheet.column_dimensions['M'].width=20
    # set headers

    headers=['id','Сана','Бригада №','Ф И О','Омбор','Юк хати № ','Транспорт маркаси','Транспорт №','Юк ташувчи','Брутто','Тара','Нетто','Юк хати жунатувчи']

    for i_col ,header in enumerate(headers,start=1):
        cell=sheet.cell(row=1,column=i_col,value=header)
        cell.font=Font(size=11,color=BLACK,bold=True)
        cell.alignment=Alignment(horizontal='center',vertical='center')


    for i_row,galla_data in enumerate(gallas_data,start=2):
        for i_col,value in enumerate(galla_data,start=1):
            cell=sheet.cell(row=i_row,column=i_col,value=value)
            cell.alignment=Alignment(horizontal='center',vertical='center')
        
        book.save(file_name)
        book.close()
        